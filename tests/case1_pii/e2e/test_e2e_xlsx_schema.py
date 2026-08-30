"""E2E test: XLSX schema validation."""
import json

import pandas as pd

from pii.export import build_document_record, export_xlsx


def _make_records():
    return [
        build_document_record(
            doc_id=f"doc-{i}",
            entities_json=[
                {"entity_id": "e1", "label": "PERSON_NAME", "start": 0, "end": 4, "score": 0.9, "needs_review": False},
            ],
            replacements={"e1": {"original": "", "pseudonym": "Test Name", "strategy": "consistent_pseudonym"}},
            anonymized_text=f"Test Name is person {i}.",
            metadata={"model": "test", "model_version": "v1", "min_score": 0.5},
            document_type="test",
            include_original=False,
        )
        for i in range(3)
    ]


class TestXlsxSchema:
    REQUIRED_COLUMNS = {"doc_id", "document_type", "entities", "replacements", "anonymized_text", "metadata"}

    def test_xlsx_has_required_columns(self, tmp_path):
        records = _make_records()
        path = tmp_path / "test.xlsx"
        export_xlsx(records, path)

        df = pd.read_excel(str(path), engine="openpyxl")
        assert self.REQUIRED_COLUMNS.issubset(set(df.columns))

    def test_xlsx_row_count(self, tmp_path):
        records = _make_records()
        path = tmp_path / "test.xlsx"
        export_xlsx(records, path)

        df = pd.read_excel(str(path), engine="openpyxl")
        assert len(df) == 3

    def test_xlsx_entities_are_valid_json(self, tmp_path):
        records = _make_records()
        path = tmp_path / "test.xlsx"
        export_xlsx(records, path)

        df = pd.read_excel(str(path), engine="openpyxl")
        for val in df["entities"]:
            parsed = json.loads(val)
            assert isinstance(parsed, list)

    def test_xlsx_replacements_are_valid_json(self, tmp_path):
        records = _make_records()
        path = tmp_path / "test.xlsx"
        export_xlsx(records, path)

        df = pd.read_excel(str(path), engine="openpyxl")
        for val in df["replacements"]:
            parsed = json.loads(val)
            assert isinstance(parsed, dict)

    def test_xlsx_opens_with_openpyxl(self, tmp_path):
        records = _make_records()
        path = tmp_path / "test.xlsx"
        export_xlsx(records, path)

        from openpyxl import load_workbook
        wb = load_workbook(str(path))
        assert wb.sheetnames == ["Sheet1"]
        ws = wb.active
        assert ws.max_row == 4  # header + 3 data rows
