import pytest

from injection_guard.config import PipelineConfig
from injection_guard.pipeline import run_pipeline

SAMPLE_SIZE = 50


@pytest.mark.model
class TestE2EFullSplitSmoke:
    def test_full_test_split_runs_without_exceptions(self, tmp_path):
        config = PipelineConfig(test_sample=SAMPLE_SIZE)
        result = run_pipeline(config=config, output_dir=tmp_path / "case2")
        assert result["rows"] == SAMPLE_SIZE
        assert result["elapsed_seconds"] > 0
        assert result["rows_per_second"] > 0

    def test_xlsx_contains_required_columns(self, tmp_path):
        import pandas as pd

        config = PipelineConfig(test_sample=SAMPLE_SIZE)
        result = run_pipeline(config=config, output_dir=tmp_path / "case2")
        df = pd.read_excel(result["xlsx_path"], engine="openpyxl")
        required_cols = {"запрос", "присвоенный класс", "рекомендуемое решение", "краткое обоснование"}
        assert required_cols.issubset(set(df.columns))
        assert len(df) == SAMPLE_SIZE

    def test_xlsx_opens_correctly(self, tmp_path):
        import openpyxl

        config = PipelineConfig(test_sample=SAMPLE_SIZE)
        result = run_pipeline(config=config, output_dir=tmp_path / "case2")
        wb = openpyxl.load_workbook(result["xlsx_path"])
        ws = wb.active
        assert ws.max_row == SAMPLE_SIZE + 1  # header + data rows
        assert ws.max_column == 4
